"""
services/excel_service.py
Complete Excel (.xlsx), CSV, and text import/export service.
Provides:
  - Excel/CSV parsing into resident records with header auto-detection
  - Duplicate property_id checking and staff merge handling
  - Sample Excel template generation for download (.xlsx)
  - Styled Tax Report Excel spreadsheet generation (.xlsx)
"""

import io
import re
import csv
import base64
import logging
from typing import Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Canonical field aliases (supports English, Marathi, common GP terms)
FIELD_MAP: dict[str, list[str]] = {
    "name": [
        "name", "taxpayer", "taxpayer name", "resident name", "owner name",
        "owner", "full name", "नाव", "करदात्याचे नाव", "मालकाचे नाव", "नांव",
    ],
    "property_id": [
        "property id", "property", "id", "prop id", "property_id",
        "property no", "property number", "property no.", "prop no",
        "house no", "house number", "house no.", "milkat no", "milkat kramank",
        "मालमत्ता क्रमांक", "घर क्र", "घर क्रमांक", "मिल्कत क्र", "मिल्कत क्रमांक",
    ],
    "ward": [
        "ward", "ward no", "ward no.", "ward number", "ward/galli",
        "प्रभाग", "वॉर्ड", "वॉर्ड क्र", "गल्ली",
    ],
    "phone": [
        "phone", "mobile", "mob", "mobile no", "mobile no.", "mobile number",
        "phone no", "phone no.", "phone number", "contact", "contact no",
        "contact no.", "contact number", "cell", "cell no",
        "मोबाईल", "फोन", "संपर्क", "मोबाईल नंबर",
    ],
    "address": [
        "address", "addr", "location", "street", "village",
        "पत्ता", "ठिकाण", "गल्ली/पत्ता",
    ],
    "base_amount": [
        "amount", "tax amount", "base amount", "tax", "base_amount",
        "tax amt", "tax amt.", "amt", "amount due", "property tax",
        "total tax", "tax due", "tax value", "tax (in rs)", "tax in rs",
        "रक्कम", "कर रक्कम", "एकूण रक्कम", "मालमत्ता कर", "कर आकारणी",
    ],
    "payment_status": [
        "status", "payment status", "payment_status", "paid/unpaid", "paid status",
        "स्थिती", "भरणा स्थिती",
    ],
    "paid_date": [
        "paid date", "payment date", "paid_date", "date paid",
        "भरणा दिनांक", "दिनांक",
    ],
}


def _normalize_key(raw_key: str) -> str:
    if not raw_key:
        return ""
    k = str(raw_key).strip().lower()
    k = re.sub(r"[.\-_/]+", " ", k)
    k = re.sub(r"\s+", " ", k).strip()
    return k


def _match_field(raw_key: str) -> str | None:
    norm = _normalize_key(raw_key)
    if not norm:
        return None
    for field, aliases in FIELD_MAP.items():
        for alias in aliases:
            if norm == _normalize_key(alias):
                return field
    return None


def _clean_amount(raw: Any) -> float:
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return 0.0
    s = re.sub(r"(?i)\b(rs|inr|rupees?)\b\.?", "", s)
    s = s.replace("₹", "").replace("/-", "").replace(",", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return 0.0
    try:
        return float(m.group(0))
    except ValueError:
        return 0.0


def _clean_phone(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    # If float string like 9876543210.0 from Excel
    if s.endswith(".0"):
        s = s[:-2]
    # Keep digits and + prefix
    cleaned = re.sub(r"[^\d+]", "", s)
    return cleaned


def parse_excel_bytes(file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse .xlsx or .xls file binary content into resident records."""
    records: list[dict[str, Any]] = []
    seen_ids = set()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        log.warning("openpyxl failed to open workbook: %s, attempting fallback", e)
        # Attempt fallback to CSV parsing if plain text
        try:
            return parse_csv_bytes(file_bytes)
        except Exception:
            raise ValueError(f"Could not read Excel file: {e}")

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # Find the header row (first row with at least 2 matching field names)
        header_row_idx = -1
        col_map: dict[int, str] = {}

        for r_idx, row in enumerate(rows[:10]):  # check top 10 rows
            matches = {}
            for c_idx, cell in enumerate(row):
                if cell is not None:
                    matched = _match_field(str(cell))
                    if matched:
                        matches[c_idx] = matched
            if "name" in matches.values() or "property_id" in matches.values():
                header_row_idx = r_idx
                col_map = matches
                break

        if header_row_idx == -1:
            # If no explicit header matched, assume standard order if enough columns
            if len(rows[0]) >= 2:
                header_row_idx = 0
                col_map = {0: "name", 1: "property_id", 2: "ward", 3: "phone", 4: "address", 5: "base_amount"}

        for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            rec: dict[str, Any] = {
                "name": "",
                "property_id": "",
                "ward": "",
                "phone": "",
                "address": "",
                "base_amount": 0.0,
                "payment_status": "unpaid",
                "paid_date": None,
                "_block_index": len(records),
                "_raw_block": f"Row {row_idx}",
            }

            for c_idx, val in enumerate(row):
                field = col_map.get(c_idx)
                if not field or val is None:
                    continue

                if field == "base_amount":
                    rec[field] = _clean_amount(val)
                elif field == "phone":
                    rec[field] = _clean_phone(val)
                elif field == "property_id":
                    rec[field] = str(val).strip()
                elif field == "payment_status":
                    v_str = str(val).strip().lower()
                    rec[field] = "paid" if "paid" in v_str or "yes" in v_str or "1" in v_str else "unpaid"
                elif field == "paid_date":
                    rec[field] = str(val).strip() if val else None
                else:
                    rec[field] = str(val).strip()

            pid = rec["property_id"].strip().lower()
            if rec["name"] and rec["property_id"] and pid not in seen_ids:
                seen_ids.add(pid)
                records.append(rec)

    return records


def parse_csv_bytes(file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse CSV content with encoding detection (UTF-8, UTF-8-BOM, Latin-1)."""
    text = ""
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue

    if not text:
        text = file_bytes.decode("utf-8", errors="replace")

    records: list[dict[str, Any]] = []
    seen_ids = set()

    # Try standard csv sniffing
    try:
        sample = text[:2048]
        dialect = csv.Sniffer().sniff(sample)
        reader = list(csv.reader(io.StringIO(text), dialect))
    except Exception:
        reader = list(csv.reader(io.StringIO(text)))

    if not reader:
        return []

    # Find header row
    header_row_idx = -1
    col_map: dict[int, str] = {}

    for r_idx, row in enumerate(reader[:10]):
        matches = {}
        for c_idx, cell in enumerate(row):
            matched = _match_field(cell)
            if matched:
                matches[c_idx] = matched
        if "name" in matches.values() or "property_id" in matches.values():
            header_row_idx = r_idx
            col_map = matches
            break

    if header_row_idx == -1:
        header_row_idx = 0
        col_map = {0: "name", 1: "property_id", 2: "ward", 3: "phone", 4: "address", 5: "base_amount"}

    for row_idx, row in enumerate(reader[header_row_idx + 1:], start=header_row_idx + 2):
        if not row or all(str(v).strip() == "" for v in row):
            continue

        rec: dict[str, Any] = {
            "name": "",
            "property_id": "",
            "ward": "",
            "phone": "",
            "address": "",
            "base_amount": 0.0,
            "payment_status": "unpaid",
            "paid_date": None,
            "_block_index": len(records),
            "_raw_block": f"CSV Row {row_idx}",
        }

        for c_idx, val in enumerate(row):
            field = col_map.get(c_idx)
            if not field or val is None:
                continue

            if field == "base_amount":
                rec[field] = _clean_amount(val)
            elif field == "phone":
                rec[field] = _clean_phone(val)
            elif field == "property_id":
                rec[field] = str(val).strip()
            elif field == "payment_status":
                v_str = str(val).strip().lower()
                rec[field] = "paid" if "paid" in v_str or "yes" in v_str or "1" in v_str else "unpaid"
            elif field == "paid_date":
                rec[field] = str(val).strip() if val else None
            else:
                rec[field] = str(val).strip()

        pid = rec["property_id"].strip().lower()
        if rec["name"] and rec["property_id"] and pid not in seen_ids:
            seen_ids.add(pid)
            records.append(rec)

    return records


def parse_text(raw_text: str) -> list[dict[str, Any]]:
    """
    Split raw pasted text into blocks (Key: Value or CSV/Tab separated).
    """
    # Check if text is tab-separated or comma-separated table
    if "\t" in raw_text or ("," in raw_text and "\n" in raw_text):
        try:
            csv_recs = parse_csv_bytes(raw_text.encode("utf-8"))
            if csv_recs:
                return csv_recs
        except Exception:
            pass

    blocks = re.split(r"\n\s*\n", raw_text.strip())
    results: list[dict[str, Any]] = []

    for block_idx, block in enumerate(blocks):
        record: dict[str, Any] = {
            "name":        "",
            "property_id": "",
            "ward":        "",
            "phone":       "",
            "address":     "",
            "base_amount": 0.0,
            "payment_status": "unpaid",
            "paid_date":   None,
            "_block_index": block_idx,
            "_raw_block":   block.strip(),
        }

        for line in block.strip().splitlines():
            line = line.strip()
            if not line:
                continue

            m = re.match(r"^(.+?)\s*[:\-]\s*(.+)$", line, re.IGNORECASE)
            if not m:
                continue

            raw_key = m.group(1).strip()
            value   = m.group(2).strip()
            matched_field = _match_field(raw_key)

            if matched_field is None:
                continue

            if matched_field == "base_amount":
                record[matched_field] = _clean_amount(value)
            elif matched_field == "phone":
                record[matched_field] = _clean_phone(value)
            else:
                record[matched_field] = value

        if record["name"] and record["property_id"]:
            results.append(record)

    return results


def check_duplicates(
    parsed_list: list[dict],
    existing_property_ids: list[str],
) -> tuple[list[dict], list[dict]]:
    """Separate parsed records into new vs duplicates."""
    existing_set = set(pid.strip().lower() for pid in existing_property_ids)
    new_records: list[dict] = []
    duplicate_records: list[dict] = []

    for rec in parsed_list:
        if rec["property_id"].strip().lower() in existing_set:
            duplicate_records.append(rec)
        else:
            new_records.append(rec)

    return new_records, duplicate_records


def confirm_import(new_records: list[dict]) -> dict:
    """Insert confirmed-new records into the DB."""
    from database.db import execute_write_many
    statements: list[tuple] = []
    errors: list[str] = []

    for rec in new_records:
        if not rec.get("name") or not rec.get("property_id"):
            errors.append(f"Skipped incomplete record: {rec.get('_raw_block', '')[:40]}")
            continue
        statements.append((
            """INSERT INTO residents
               (name, property_id, ward, phone, address, base_amount, payment_status, paid_date)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                rec["name"],
                rec["property_id"],
                rec.get("ward", ""),
                rec.get("phone", ""),
                rec.get("address", ""),
                rec.get("base_amount", 0.0),
                rec.get("payment_status", "unpaid"),
                rec.get("paid_date", None),
            )
        ))

    if statements:
        execute_write_many(statements)

    return {"inserted": len(statements), "errors": errors}


def merge_resident(existing_id: int, new_data: dict) -> dict:
    """Staff-approved overwrite of an existing resident's mutable fields."""
    from database.db import execute_write
    execute_write(
        """UPDATE residents
           SET name=?, ward=?, phone=?, address=?, base_amount=?
           WHERE id=?""",
        (
            new_data.get("name", ""),
            new_data.get("ward", ""),
            new_data.get("phone", ""),
            new_data.get("address", ""),
            new_data.get("base_amount", 0.0),
            existing_id,
        )
    )
    return {"merged": True, "id": existing_id}


# ── Excel Template Generator ──────────────────────────────────────────────────

def generate_sample_excel_template() -> bytes:
    """
    Generates a pre-formatted, styled Excel sample template (.xlsx)
    ready for users to fill in taxpayer records.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Taxpayer Import Template"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    header_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    sample_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    headers = [
        "Taxpayer Name *",
        "Property ID *",
        "Ward",
        "Phone Number *",
        "Address",
        "Base Tax Amount (₹) *",
        "Payment Status (unpaid/paid)",
    ]

    sample_rows = [
        ["Ramesh Patil", "GP/2026/001", "Ward 1", "9876543210", "Plot 12, Main Road, Gram Panchayat", 2500, "unpaid"],
        ["Sunita Deshmukh", "GP/2026/002", "Ward 2", "9123456789", "House 45, Shivaji Nagar", 1800, "unpaid"],
        ["Vijay Shinde", "GP/2026/003", "Ward 1", "9988776655", "Gat No 7, Near Gram Panchayat Temple", 3200, "paid"],
        ["Anil Gaikwad", "GP/2026/004", "Ward 3", "9822114433", "Shop 3, Market Yard", 4500, "unpaid"],
        ["सुनील शिंदे", "GP/2026/005", "Ward 2", "9766554433", "घर क्र १२, मारुती मंदिर जवळ", 2100, "unpaid"],
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.row_dimensions[1].height = 28

    for r_idx, row_data in enumerate(sample_rows, 2):
        ws.row_dimensions[r_idx].height = 20
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font = sample_font
            cell.border = thin_border
            if col_idx in (2, 4, 7):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column auto widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ── Excel Report Generator ────────────────────────────────────────────────────

def generate_excel_report(
    report_dict: dict,
    period_label: str = "All Records",
    gp_name: str = "Gram Panchayat Office",
) -> bytes:
    """
    Generates a professionally styled Property Tax Statement Excel (.xlsx) file.
    Includes title banner, summary KPIs, detailed taxpayer table, and totals.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tax Report Statement"
    ws.views.sheetView[0].showGridLines = True

    # Color definitions
    c_brand = "D97706"      # Saffron Amber
    c_dark = "1C1409"       # Deep Brown
    c_cream = "FEF3C7"      # Saffron light
    c_green_fill = "F0FDF4"
    c_green_text = "065F46"
    c_red_fill = "FEF2F2"
    c_red_text = "991B1B"

    # Fonts
    f_title = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    f_sub = Font(name="Arial", size=10, italic=True, color="FFFFFF")
    f_kpi_num = Font(name="Arial", size=13, bold=True)
    f_kpi_lbl = Font(name="Arial", size=9, bold=True, color="5C4030")
    f_tbl_hdr = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    f_data = Font(name="Arial", size=10)
    f_data_bold = Font(name="Arial", size=10, bold=True)
    f_paid = Font(name="Arial", size=10, bold=True, color=c_green_text)
    f_unpaid = Font(name="Arial", size=10, bold=True, color=c_red_text)

    # Fills & Borders
    fill_header = PatternFill(start_color=c_brand, end_color=c_brand, fill_type="solid")
    fill_title = PatternFill(start_color=c_dark, end_color=c_dark, fill_type="solid")
    fill_kpi_paid = PatternFill(start_color=c_green_fill, end_color=c_green_fill, fill_type="solid")
    fill_kpi_unpaid = PatternFill(start_color=c_red_fill, end_color=c_red_fill, fill_type="solid")
    fill_kpi_tot = PatternFill(start_color=c_cream, end_color=c_cream, fill_type="solid")
    fill_alt = PatternFill(start_color="FFFBF5", end_color="FFFBF5", fill_type="solid")

    border_thin = Border(
        left=Side(style="thin", color="EDE5D8"),
        right=Side(style="thin", color="EDE5D8"),
        top=Side(style="thin", color="EDE5D8"),
        bottom=Side(style="thin", color="EDE5D8"),
    )
    border_total = Border(
        top=Side(style="thin", color="1C1409"),
        bottom=Side(style="double", color="1C1409"),
    )

    # 1. Main Title Banner (Row 1-2)
    ws.merge_cells("A1:K1")
    cell_title = ws["A1"]
    cell_title.value = f"🏛 {gp_name} — Property Tax Statement"
    cell_title.font = f_title
    cell_title.fill = fill_title
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    cell_sub = ws["A2"]
    from datetime import datetime
    now_str = datetime.now().strftime("%d-%b-%Y %I:%M %p")
    cell_sub.value = f"Report Scope: {period_label}  |  Generated on: {now_str}  |  Home Tax Reminder System"
    cell_sub.font = f_sub
    cell_sub.fill = fill_title
    cell_sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 2. KPI Summary Boxes (Row 4-5)
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 24

    # Paid Collected Box (Cols B-C)
    ws.merge_cells("B4:C4")
    ws["B4"].value = "✓ TAX COLLECTED (PAID)"
    ws["B4"].font = f_kpi_lbl
    ws["B4"].fill = fill_kpi_paid
    ws["B4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B5:C5")
    ws["B5"].value = report_dict.get("paid_sum", 0.0)
    ws["B5"].font = Font(name="Arial", size=12, bold=True, color=c_green_text)
    ws["B5"].number_format = "₹#,##0.00"
    ws["B5"].fill = fill_kpi_paid
    ws["B5"].alignment = Alignment(horizontal="center", vertical="center")

    # Unpaid Outstanding Box (Cols E-F)
    ws.merge_cells("E4:F4")
    ws["E4"].value = "⚠ OUTSTANDING DUES (UNPAID)"
    ws["E4"].font = f_kpi_lbl
    ws["E4"].fill = fill_kpi_unpaid
    ws["E4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("E5:F5")
    ws["E5"].value = report_dict.get("unpaid_sum", 0.0)
    ws["E5"].font = Font(name="Arial", size=12, bold=True, color=c_red_text)
    ws["E5"].number_format = "₹#,##0.00"
    ws["E5"].fill = fill_kpi_unpaid
    ws["E5"].alignment = Alignment(horizontal="center", vertical="center")

    # Total Taxable Value Box (Cols H-I)
    ws.merge_cells("H4:I4")
    ws["H4"].value = "📊 TOTAL TAX VALUE"
    ws["H4"].font = f_kpi_lbl
    ws["H4"].fill = fill_kpi_tot
    ws["H4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("H5:I5")
    ws["H5"].value = float(report_dict.get("paid_sum", 0.0) + report_dict.get("unpaid_sum", 0.0))
    ws["H5"].font = Font(name="Arial", size=12, bold=True, color="92400E")
    ws["H5"].number_format = "₹#,##0.00"
    ws["H5"].fill = fill_kpi_tot
    ws["H5"].alignment = Alignment(horizontal="center", vertical="center")

    # 3. Data Table Headers (Row 7)
    table_headers = [
        "Property ID",
        "Taxpayer Name",
        "Ward",
        "Phone Number",
        "Address",
        "Payment Status",
        "Base Tax (₹)",
        "Penalty (₹)",
        "Rebate (₹)",
        "Net Amount (₹)",
        "Date (Paid / Due Date)",
    ]

    ws.row_dimensions[7].height = 26
    for col_idx, h in enumerate(table_headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=h)
        cell.font = f_tbl_hdr
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # 4. Data Rows
    records = report_dict.get("records", [])
    start_row = 8
    for idx, r in enumerate(records):
        cur_row = start_row + idx
        ws.row_dimensions[cur_row].height = 20
        is_paid = r.get("payment_status") == "paid"
        date_display = (r.get("paid_date") if is_paid else r.get("due_date")) or "—"
        row_fill = fill_alt if idx % 2 == 1 else PatternFill(fill_type=None)

        row_vals = [
            r.get("property_id", ""),
            r.get("name", ""),
            r.get("ward", "") or "—",
            r.get("phone", "") or "—",
            r.get("address", "") or "—",
            "PAID" if is_paid else "UNPAID",
            r.get("base_amount", 0.0),
            r.get("penalty", 0.0),
            r.get("rebate", 0.0),
            r.get("net_due", r.get("base_amount", 0.0)),
            date_display,
        ]

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=cur_row, column=col_idx, value=val)
            cell.border = border_thin
            if row_fill.fill_type:
                cell.fill = row_fill

            if col_idx == 1:
                cell.font = f_data_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 6:
                cell.font = f_paid if is_paid else f_unpaid
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (7, 8, 9, 10):
                cell.font = f_data_bold if col_idx == 10 else f_data
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.00"
            elif col_idx in (3, 4, 11):
                cell.font = f_data
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = f_data
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 5. Total Row
    tot_row = start_row + len(records)
    if records:
        ws.row_dimensions[tot_row].height = 24
        ws.cell(row=tot_row, column=1, value="Total Summary:").font = f_data_bold
        ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="right", vertical="center")

        ws.cell(row=tot_row, column=6, value=f"{len(records)} Records").font = f_data_bold
        ws.cell(row=tot_row, column=6).alignment = Alignment(horizontal="center", vertical="center")

        # Formulas or direct sums
        ws.cell(row=tot_row, column=7, value=f"=SUM(G{start_row}:G{tot_row-1})").font = f_data_bold
        ws.cell(row=tot_row, column=7).number_format = "#,##0.00"
        ws.cell(row=tot_row, column=7).alignment = Alignment(horizontal="right", vertical="center")

        ws.cell(row=tot_row, column=8, value=f"=SUM(H{start_row}:H{tot_row-1})").font = f_data_bold
        ws.cell(row=tot_row, column=8).number_format = "#,##0.00"
        ws.cell(row=tot_row, column=8).alignment = Alignment(horizontal="right", vertical="center")

        ws.cell(row=tot_row, column=9, value=f"=SUM(I{start_row}:I{tot_row-1})").font = f_data_bold
        ws.cell(row=tot_row, column=9).number_format = "#,##0.00"
        ws.cell(row=tot_row, column=9).alignment = Alignment(horizontal="right", vertical="center")

        ws.cell(row=tot_row, column=10, value=f"=SUM(J{start_row}:J{tot_row-1})").font = f_data_bold
        ws.cell(row=tot_row, column=10).number_format = "#,##0.00"
        ws.cell(row=tot_row, column=10).alignment = Alignment(horizontal="right", vertical="center")

        for c in range(1, 12):
            ws.cell(row=tot_row, column=c).border = border_total

    # Column Auto-Widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            # Skip title banner rows from length calculation
            if cell.row in (1, 2, 4, 5):
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
